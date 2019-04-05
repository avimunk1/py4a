#this progrem  fix phone numbers

def fixNumeric(phone):
    print('start',phone)
    phone1=''
    for char in phone:
        print('inloop',char)
        if not char.isnumeric():
            continue
        else:
            phone1=phone1+char
            print('else' ,char,phone1)
    return(phone1)

def fixPrefix(phone):
    prefixstatus=0
    print(phone)
    if phone.startswith('5'):
        phone=phone='0'+phone
        return(phone)
    if phone.startswith('972'):
       phone=phone[3:]
       phone=phone='0'+phone
       return(phone)
    else:
       return('error')

def checkMobilePhone(phone):
    checkStatus='0'
    phone1='9'+phone
    if not len(phone)==10:
        checkStatus='1'
    if not phone.startswith('05'):
        checkStatus='2'
    if not phone1.isnumeric():
        checkStatus='3'
    return(checkStatus)

#file hendling
inputFile="input_sample.csv"
outPutFile="fixedCustomersList.csv"

try:
    fileH=open('inputs/'+inputFile)
except:
    print('no such file'+inputFile)
    exit()

fileH=open('inputs/'+inputFile)
outFilH=open(outPutFile,"w+")

#create the list
fileLine=list()
ExtId=''
FirstName=''
LastName=''
MobilePhone=''
AltPhone=''
newMobilePhone=''
newPhoneSatus=''
outFilH=open(outPutFile,"w+")
from openpyxl import Workbook
wb = Workbook()
excelLine=1
excelColmn=1
ws = wb.active
for line in fileH:
    fileLine=line.split(',')
    ExtId=fileLine[0]
    FirstName=fileLine[1]
    LastName=fileLine[2]
    MobilePhone=fileLine[3]
    newMobilePhone=fixNumeric(MobilePhone)
    newMobilePhone=fixPrefix(newMobilePhone)
    newPhoneSatus=checkMobilePhone(newMobilePhone)

    d = ws.cell(row=excelLine, column=excelColmn, value=ExtId)
    excelColmn=excelColmn+1
    d = ws.cell(row=excelLine, column=excelColmn, value=FirstName)
    excelColmn=excelColmn+1
    d = ws.cell(row=excelLine, column=excelColmn, value=LastName)
    excelColmn=excelColmn+1
    d = ws.cell(row=excelLine, column=excelColmn, value=MobilePhone)
    excelColmn=excelColmn+1
    d = ws.cell(row=excelLine, column=excelColmn, value=newMobilePhone)
    excelColmn=excelColmn+1
    d = ws.cell(row=excelLine, column=excelColmn, value=newPhoneSatus)
    excelColmn=excelColmn+1
    excelLine=excelLine+1
    excelColmn=1
wb.save('avitemp.xlsx')
